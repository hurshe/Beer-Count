"""
export_excel.py — Export Beer Count to Excel
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import database as db

# ── Palette ───────────────────────────────────────
GOLD    = "C9933A"; GOLD_LT = "FDF3DF"; HDR_BG  = "3A3020"
GREEN   = "2A7A45"; GREEN_LT= "EAF6EE"
RED     = "B83232"; RED_LT  = "FDEAEA"
AMBER   = "A06010"; AMBER_LT= "FFF4E0"
ORANGE  = "C05000"; ORANGE_LT="FFF0E0"
BORDER  = "D0C8BA"; WHITE   = "FFFFFF"; GREY    = "F5F2EC"

def _side(): return Side(border_style="thin", color=BORDER)
def _border(): return Border(left=_side(), right=_side(), top=_side(), bottom=_side())
def _fill(c): return PatternFill("solid", fgColor=c)

def _hdr(cell, color=WHITE, bold=True, size=9):
    cell.font = Font(bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()

def _auto_width(ws, max_w=24):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, max_w)

def _diff_fill(diff):
    s = db.diff_status(diff)
    return {"ok": GREEN_LT, "warn": AMBER_LT, "over": ORANGE_LT, "bad": RED_LT}[s]

def _diff_color(diff):
    s = db.diff_status(diff)
    return {"ok": GREEN, "warn": AMBER, "over": ORANGE, "bad": RED}[s]

def _diff_label(diff):
    s = db.diff_status(diff)
    icons = {"ok":"✅","warn":"🟡","over":"🟠","bad":"🔴"}
    return f"{icons[s]} {diff:+.2f}L"


# ── Month export ──────────────────────────────────
def export_month(month: str, filepath: str) -> bool:
    entries = db.get_days_for_month(month)
    if not entries: return False
    sizes = db.get_sizes()  # fetch once for all day sheets
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("📊 Podsumowanie")
    _write_summary(ws, month, entries)
    for entry_date, data in entries:
        ws2 = wb.create_sheet(f"Dzień {entry_date[8:]}")
        _write_day(ws2, entry_date, data, sizes)
    wb.save(filepath)
    return True


def export_year(year: str, filepath: str) -> bool:
    all_days = [(d, data) for d, data in db.get_all_days() if d.startswith(year)]
    if not all_days: return False
    sizes = db.get_sizes()  # fetch once for all day sheets
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws_ann = wb.create_sheet(f"📅 Rok {year}")
    _write_annual(ws_ann, year, all_days)
    months = sorted(set(d[:7] for d, _ in all_days))
    for month in months:
        entries = [(d, data) for d, data in all_days if d.startswith(month)]
        ws = wb.create_sheet(_month_name(month)[:31])
        _write_summary(ws, month, entries)
    wb.save(filepath)
    return True


# ── Sheet writers ─────────────────────────────────
def _write_day(ws, entry_date, data, sizes):
    beers = data.get("kegs", [])

    ws["A1"] = f"Beer Count — {entry_date}"
    ws["A1"].font = Font(bold=True, size=13, color=GOLD)
    ws["A1"].fill = _fill(HDR_BG)
    ws.merge_cells(f"A1:N1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # KEGS section
    row = 3
    ws.cell(row=row, column=1, value="STAN KEGÓW").font = Font(bold=True, size=10, color=GOLD)
    row += 1
    hdrs = ["Piwo", "Keg", "START (L)", "Dostawa (szt.)", "PEŁNE END",
            "Otw.kg#1", "Otw.kg#2", "Otw.kg#3", "END (L)", "POS (L)", "Korekty (L)", "RÓŻNICA"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = _fill(GOLD_LT); _hdr(c, color=GOLD)
    row += 1

    for i, keg in enumerate(beers):
        pos_entry = data["pos"][i] if i < len(data.get("pos", [])) else {"sizes": []}
        corr_data = data["corr"][i] if i < len(data.get("corr", [])) else {}
        end_l = db.keg_end_liters(keg)
        pos_l = db.pos_liters(pos_entry)
        corr_l = db.corr_liters(corr_data)
        diff  = db.calc_diff(keg, pos_entry, corr_data)
        ow = keg.get("open_end") or [None, None, None]

        vals = [
            keg["name"], f"{keg['keg']}L",
            round(db.keg_start_liters(keg), 2),
            keg.get("delivery", 0),
            keg.get("full_end", 0),
            ow[0] if len(ow) > 0 else "",
            ow[1] if len(ow) > 1 else "",
            ow[2] if len(ow) > 2 else "",
            round(end_l, 2), round(pos_l, 2), round(corr_l, 2),
            _diff_label(diff),
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = _border()
            c.alignment = Alignment(horizontal="center")
            if ci == 1: c.font = Font(bold=True); c.alignment = Alignment(horizontal="left")
            if ci == 12:
                c.fill = _fill(_diff_fill(diff))
                c.font = Font(bold=True, color=_diff_color(diff))
        row += 1

    # Legend
    row += 1
    legends = [("✅ ±2L — Norma", GREEN), ("🟡 +2/+5L — Sprawdź POS", AMBER),
               ("🟠 >+5L — Sprzedaż poza POS?", ORANGE), ("🔴 <−5L — Straty/spille", RED)]
    for ci, (txt, col) in enumerate(legends, 1):
        c = ws.cell(row=row, column=ci, value=txt)
        c.font = Font(color=col, bold=True, size=8)

    _auto_width(ws)


def _write_summary(ws, month, entries):
    beers_cfg = db.get_beers()
    beer_names = [b["name"] for b in beers_cfg]

    ws["A1"] = f"Raport — {_month_name(month)}"
    ws["A1"].font = Font(bold=True, size=13, color=GOLD)
    ws["A1"].fill = _fill(HDR_BG)
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="center")

    # KPIs
    row = 3
    tot_days = len(entries)
    tot_del  = sum(sum(int(k.get("delivery",0) or 0) for k in d.get("kegs",[])) for _,d in entries)
    tot_pos  = sum(sum(db.pos_liters(p) for p in d.get("pos",[])) for _,d in entries)
    tot_corr = sum(sum(db.corr_liters(c) for c in d.get("corr",[])) for _,d in entries)

    # total diff across all
    tot_diff = 0
    for _, data in entries:
        for i, keg in enumerate(data.get("kegs",[])):
            pe = data["pos"][i] if i<len(data.get("pos",[])) else {"sizes":[]}
            co = data["corr"][i] if i<len(data.get("corr",[])) else {}
            tot_diff += db.calc_diff(keg, pe, co)

    kpis = [("Dni wpisów", tot_days), ("Dostawa (kegi)", tot_del),
            ("Sprzedaż POS (L)", round(tot_pos,1)), ("Korekty (L)", round(tot_corr,1)),
            ("Różnica (L)", round(tot_diff,1))]
    for ci, (lbl, val) in enumerate(kpis, 1):
        ws.cell(row=row, column=ci, value=lbl).font = Font(bold=True, size=8, color=GOLD)
        ws.cell(row=row, column=ci).fill = _fill(GOLD_LT)
        vc = ws.cell(row=row+1, column=ci, value=val)
        vc.font = Font(bold=True, size=14,
                       color=(_diff_color(val) if lbl=="Różnica (L)" else GOLD))
        vc.alignment = Alignment(horizontal="center")

    # Per-beer table
    row = 7
    ws.cell(row=row, column=1, value="Wynik per piwo").font = Font(bold=True, size=10, color=GOLD)
    row += 1
    for ci, h in enumerate(["Piwo","Zużycie (L)","POS (L)","Korekty (L)","Różnica (L)","Dostawa (kegi)","Dni"], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = _fill(GOLD_LT); _hdr(c, color=GOLD)
    row += 1

    agg = {n:{"zuz":0,"pos":0,"corr":0,"diff":0,"del":0,"days":0} for n in beer_names}
    for _, data in entries:
        for i, keg in enumerate(data.get("kegs",[])):
            n = keg["name"]
            if n not in agg: agg[n]={"zuz":0,"pos":0,"corr":0,"diff":0,"del":0,"days":0}
            pe = data["pos"][i] if i<len(data.get("pos",[])) else {"sizes":[]}
            co = data["corr"][i] if i<len(data.get("corr",[])) else {}
            end_l = db.keg_end_liters(keg)
            start_l = db.keg_start_liters(keg)
            del_l = db.keg_delivery_liters(keg)
            zuz = start_l + del_l - end_l
            agg[n]["zuz"]  += zuz
            agg[n]["pos"]  += db.pos_liters(pe)
            agg[n]["corr"] += db.corr_liters(co)
            agg[n]["diff"] += db.calc_diff(keg, pe, co)
            agg[n]["del"]  += int(keg.get("delivery",0) or 0)
            agg[n]["days"] += 1

    for n, a in agg.items():
        diff = round(a["diff"], 2)
        vals = [n, round(a["zuz"],1), round(a["pos"],1), round(a["corr"],1),
                _diff_label(diff), a["del"], a["days"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = _border(); c.alignment = Alignment(horizontal="center")
            if ci==1: c.font=Font(bold=True); c.alignment=Alignment(horizontal="left")
            if ci==5: c.fill=_fill(_diff_fill(diff)); c.font=Font(bold=True, color=_diff_color(diff))
        row += 1

    # Daily trend
    row += 1
    ws.cell(row=row, column=1, value="Trend dzienny").font = Font(bold=True, size=10, color=GOLD)
    row += 1
    for ci, h in enumerate(["Data"] + beer_names + ["Łącznie diff (L)"], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = _fill(GOLD_LT); _hdr(c, color=GOLD)

    for entry_date, data in entries:
        row += 1
        ws.cell(row=row, column=1, value=entry_date).font = Font(bold=True)
        day_diff = 0
        for i, keg in enumerate(data.get("kegs",[])):
            n = keg["name"]
            pe = data["pos"][i] if i<len(data.get("pos",[])) else {"sizes":[]}
            co = data["corr"][i] if i<len(data.get("corr",[])) else {}
            diff = db.calc_diff(keg, pe, co)
            day_diff += diff
            if n in beer_names:
                ci = beer_names.index(n) + 2
                c = ws.cell(row=row, column=ci, value=round(diff,2))
                c.fill = _fill(_diff_fill(diff))
                c.font = Font(color=_diff_color(diff), bold=True)
                c.border = _border(); c.alignment = Alignment(horizontal="center")
        dc = ws.cell(row=row, column=len(beer_names)+2, value=round(day_diff,2))
        dc.fill = _fill(_diff_fill(day_diff))
        dc.font = Font(bold=True, color=_diff_color(day_diff))
        dc.border = _border(); dc.alignment = Alignment(horizontal="center")

    _auto_width(ws)


def _write_annual(ws, year, year_days):
    ws["A1"] = f"Raport roczny — {year}"
    ws["A1"].font = Font(bold=True, size=13, color=GOLD)
    ws["A1"].fill = _fill(HDR_BG)
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="center")
    row = 3
    for ci, h in enumerate(["Miesiąc","Dni","Dostawa (kegi)","POS (L)","Korekty (L)","Różnica (L)"], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = _fill(GOLD_LT); _hdr(c, color=GOLD)
    row += 1
    months = sorted(set(d[:7] for d, _ in year_days))
    for month in months:
        entries = [(d, data) for d, data in year_days if d.startswith(month)]
        days  = len(entries)
        deliv = sum(sum(int(k.get("delivery",0) or 0) for k in d.get("kegs",[])) for _,d in entries)
        pos   = sum(sum(db.pos_liters(p) for p in d.get("pos",[])) for _,d in entries)
        corr  = sum(sum(db.corr_liters(c) for c in d.get("corr",[])) for _,d in entries)
        diff  = sum(db.calc_diff(k, d["pos"][i] if i<len(d.get("pos",[])) else {"sizes":[]},
                                  d["corr"][i] if i<len(d.get("corr",[])) else {})
                    for _,d in entries for i,k in enumerate(d.get("kegs",[])))
        diff  = round(diff, 2)
        vals = [_month_name(month), days, deliv, round(pos,1), round(corr,1), _diff_label(diff)]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = _border(); c.alignment = Alignment(horizontal="center")
            if ci==1: c.font=Font(bold=True); c.alignment=Alignment(horizontal="left")
            if ci==6: c.fill=_fill(_diff_fill(diff)); c.font=Font(bold=True, color=_diff_color(diff))
        row += 1
    _auto_width(ws)


def _month_name(m):
    MO = {"01":"Styczeń","02":"Luty","03":"Marzec","04":"Kwiecień","05":"Maj","06":"Czerwiec",
          "07":"Lipiec","08":"Sierpień","09":"Wrzesień","10":"Październik","11":"Listopad","12":"Grudzień"}
    y, mo = m.split("-")
    return f"{MO.get(mo, mo)} {y}"
